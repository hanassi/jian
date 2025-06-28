import os
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# XML 폴더 위치
xml_folder = "./"
# 출력 엑셀 파일명
output_excel = "merged_results.xlsx"

# 새 워크북 생성
wb = Workbook()
# 기본 생성된 시트 제거
wb.remove(wb.active)

# xml 폴더 내 모든 XML 파일 처리
for filename in os.listdir(xml_folder):
    if filename.endswith(".xml"):
        file_path = os.path.join(xml_folder, filename)
        tree = ET.parse(file_path)
        root = tree.getroot()

        # 새 시트 생성 (파일명 기준, 31자 제한)
        sheet_name = os.path.splitext(filename)[0][:31]
        ws = wb.create_sheet(title=sheet_name)

        # 헤더 작성
        headers = ["항목번호", "제목", "결과", "내용", "참고"]
        ws.append(headers)

        # 각 item 요소 처리
        for item in root.findall("item"):
            row = [
                item.findtext("item_num", "").strip(),
                item.findtext("item_title", "").strip(),
                item.findtext("item_result", "").strip(),
                item.findtext("contents", "").strip(),
                item.findtext("ref", "").strip()
            ]
            ws.append(row)

        # 열 너비 자동 조정 (간단하게 최대 길이 기반)
        for col_num, col_cells in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width

# 엑셀 파일 저장
wb.save(output_excel)
