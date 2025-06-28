import openpyxl
from glob import glob

# 원본 파일(.xlsm) 경로
path = "./"

# 새로 저장할 파일(.xlsx) 경로
target_file = "보안 가이드.xlsx"

# 최종 워크북 생성
wb_target = openpyxl.Workbook()
wb_target.remove(wb_target.active)

# 시트 이름에 포함된 키워드
target_keyword = "#숨김. "
guide_keyword = "가이드"

# 원본 파일 열기
filelist = glob(f"{path}/*.xlsm")
for file in filelist:
    wb_source = openpyxl.load_workbook(file, data_only=True)

    # 대상 시트 찾기
    target_sheet = None
    for sheet_name in wb_source.sheetnames:
        if target_keyword in sheet_name and guide_keyword in sheet_name:
            target_sheet = wb_source[sheet_name]
            break

    if target_sheet is None:
        print("대상 시트를 찾을 수 없습니다.")
    else:
        # 새 워크시트 생성
        ws_target = wb_target.create_sheet(title=target_sheet.title.strip(target_keyword))

        # 셀 내용 복사
        for row in target_sheet.iter_rows():
            for cell in row:
                ws_target[cell.coordinate].value = cell.value

    # 새 파일 저장
    wb_target.save(target_file)
    print(f"시트 '{target_sheet.title}'이(가) '{target_file}'로 복사되었습니다.")
